import os
import json
from http.server import HTTPServer, SimpleHTTPRequestHandler
from openpyxl import Workbook
from io import BytesIO

from database.init_db import init_db
from database.connection import get_connection

# Inicializa la base de datos
init_db()

def generate_excel():
    # Get all data
    usuarios = _db_query_static("SELECT id, nombre, email, rol FROM usuarios ORDER BY id ASC")
    salas = _db_query_static("SELECT id, nombre, capacidad, descripcion FROM salas ORDER BY id ASC")
    reservas = _db_query_static("SELECT id, fecha, hora_inicio, hora_fin, usuario_id, sala_id FROM reservas ORDER BY id ASC")

    # Create workbook
    wb = Workbook()

    # Usuarios sheet
    ws_usuarios = wb.active
    ws_usuarios.title = "Usuarios"
    ws_usuarios.append(["ID", "Nombre", "Email", "Rol"])
    for row in usuarios:
        ws_usuarios.append(row)

    # Salas sheet
    ws_salas = wb.create_sheet("Salas")
    ws_salas.append(["ID", "Nombre", "Capacidad", "Descripción"])
    for row in salas:
        ws_salas.append(row)

    # Reservas sheet
    ws_reservas = wb.create_sheet("Reservas")
    ws_reservas.append(["ID", "Fecha", "Hora Inicio", "Hora Fin", "Usuario ID", "Sala ID"])
    for row in reservas:
        ws_reservas.append(row)

    # Save to file
    wb.save("frontend/coworking_data.xlsx")
    print("Excel actualizado")

# Helper function for static query
def _db_query_static(query, params=()):
    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        return rows
    finally:
        conn.close()

print("Backend inicializado correctamente.")
# Generate initial Excel
generate_excel()

class RequestHandler(SimpleHTTPRequestHandler):
    def _send_json(self, status_code: int, payload):
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status_code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_excel(self, workbook):
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        body = buffer.getvalue()
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", "attachment; filename=coworking_data.xlsx")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _read_json_body(self):
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length) if length > 0 else b""
        if not raw:
            return None
        try:
            return json.loads(raw.decode("utf-8"))
        except json.JSONDecodeError:
            return None

    def _path_parts(self):
        # Retorna partes de la ruta, p.ej. /api/usuarios/1 -> ['api','usuarios','1']
        return [p for p in self.path.split('?')[0].split('/') if p]

    # Utilidades de acceso a BD simples y directas
    def _db_query(self, query, params=()):
        conn = get_connection()
        try:
            cur = conn.cursor()
            cur.execute(query, params)
            rows = cur.fetchall()
            return rows
        finally:
            conn.close()

    def _db_execute(self, query, params=()):
        conn = get_connection()
        try:
            cur = conn.cursor()
            cur.execute(query, params)
            conn.commit()
            return cur.lastrowid
        finally:
            conn.close()

    def _db_execute_rowcount(self, query, params=()):
        conn = get_connection()
        try:
            cur = conn.cursor()
            cur.execute(query, params)
            conn.commit()
            return cur.rowcount
        finally:
            conn.close()

    def do_GET(self):
        if self.path == "/api/usuarios":
            rows = self._db_query("SELECT id, nombre, email, rol FROM usuarios ORDER BY id ASC")
            usuarios = [
                {"id": r[0], "nombre": r[1], "email": r[2], "rol": r[3]}
                for r in rows
            ]
            return self._send_json(200, usuarios)

        if self.path == "/api/salas":
            rows = self._db_query("SELECT id, nombre, capacidad, descripcion FROM salas ORDER BY id ASC")
            salas = [
                {"id": r[0], "nombre": r[1], "capacidad": r[2], "descripcion": r[3]}
                for r in rows
            ]
            return self._send_json(200, salas)

        if self.path == "/api/reservas":
            rows = self._db_query(
                "SELECT id, fecha, hora_inicio, hora_fin, usuario_id, sala_id FROM reservas ORDER BY id ASC"
            )
            reservas = [
                {
                    "id": r[0],
                    "fecha": r[1],
                    "hora_inicio": r[2],
                    "hora_fin": r[3],
                    "usuario_id": r[4],
                    "sala_id": r[5],
                }
                for r in rows
            ]
            return self._send_json(200, reservas)

        if self.path == "/coworking_data.xlsx":
            try:
                with open("coworking_data.xlsx", "rb") as f:
                    self.send_response(200)
                    self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    self.send_header("Content-Disposition", "attachment; filename=coworking_data.xlsx")
                    self.end_headers()
                    self.wfile.write(f.read())
            except FileNotFoundError:
                self.send_error(404, "File not found")
            return

        # Para todo lo demás, servir archivos estáticos del directorio actual
        return super().do_GET()

    def do_POST(self):
        # Usuarios
        if self.path == "/api/usuarios":
            data = self._read_json_body() or {}
            nombre = data.get("nombre")
            email = data.get("email")
            rol = data.get("rol")
            if not nombre or not email or not rol:
                return self._send_json(400, {"error": "Campos requeridos: nombre, email, rol"})

            try:
                new_id = self._db_execute(
                    "INSERT INTO usuarios (nombre, email, rol) VALUES (?, ?, ?)",
                    (nombre, email, rol),
                )
            except Exception as e:
                return self._send_json(400, {"error": str(e)})

            generate_excel()
            return self._send_json(201, {"id": new_id, "nombre": nombre, "email": email, "rol": rol})

        # Salas
        if self.path == "/api/salas":
            data = self._read_json_body() or {}
            nombre = data.get("nombre")
            capacidad = data.get("capacidad")
            descripcion = data.get("descripcion")
            if not nombre or capacidad in (None, ""):
                return self._send_json(400, {"error": "Campos requeridos: nombre, capacidad"})

            try:
                cap_int = int(capacidad)
            except ValueError:
                return self._send_json(400, {"error": "capacidad debe ser entero"})

            new_id = self._db_execute(
                "INSERT INTO salas (nombre, capacidad, descripcion) VALUES (?, ?, ?)",
                (nombre, cap_int, descripcion),
            )
            generate_excel()
            return self._send_json(201, {"id": new_id, "nombre": nombre, "capacidad": cap_int, "descripcion": descripcion})

        # Reservas
        if self.path == "/api/reservas":
            data = self._read_json_body() or {}
            fecha = data.get("fecha")  # "YYYY-MM-DD"
            hora_inicio = data.get("hora_inicio")  # "HH:MM"
            hora_fin = data.get("hora_fin")  # "HH:MM"
            usuario_id = data.get("usuario_id")
            sala_id = data.get("sala_id")

            required = (fecha, hora_inicio, hora_fin, usuario_id, sala_id)
            if any(v in (None, "") for v in required):
                return self._send_json(400, {"error": "Campos requeridos: fecha, hora_inicio, hora_fin, usuario_id, sala_id"})

            # Normalizar tipos: IDs deben ser enteros
            try:
                usuario_id = int(usuario_id)
                sala_id = int(sala_id)
            except (TypeError, ValueError):
                return self._send_json(400, {"error": "usuario_id y sala_id deben ser enteros"})

            # Validar existencia de usuario y sala
            u = self._db_query("SELECT 1 FROM usuarios WHERE id = ?", (usuario_id,))
            if not u:
                return self._send_json(404, {"error": "El usuario no existe"})
            s = self._db_query("SELECT 1 FROM salas WHERE id = ?", (sala_id,))
            if not s:
                return self._send_json(404, {"error": "La sala no existe"})

            # Validación de superposición en la misma sala y fecha
            overlaps = self._db_query(
                "SELECT hora_inicio, hora_fin FROM reservas WHERE sala_id = ? AND fecha = ?",
                (sala_id, fecha),
            )
            for hi, hf in overlaps:
                # Asumiendo formato HH:MM lexicográficamente comparable
                if (hora_inicio < hf) and (hora_fin > hi):
                    return self._send_json(409, {"error": "La sala ya está reservada en ese horario."})

            try:
                new_id = self._db_execute(
                    "INSERT INTO reservas (fecha, hora_inicio, hora_fin, usuario_id, sala_id) VALUES (?, ?, ?, ?, ?)",
                    (fecha, hora_inicio, hora_fin, usuario_id, sala_id),
                )
            except Exception as e:
                return self._send_json(400, {"error": str(e)})

            generate_excel()
            return self._send_json(
                201,
                {
                    "id": new_id,
                    "fecha": fecha,
                    "hora_inicio": hora_inicio,
                    "hora_fin": hora_fin,
                    "usuario_id": usuario_id,
                    "sala_id": sala_id,
                },
            )

        # Si no es una ruta API conocida
        return self._send_json(404, {"error": "Ruta no encontrada"})

    def do_DELETE(self):
        parts = self._path_parts()
        # Espera /api/<recurso>/<id>
        if len(parts) == 3 and parts[0] == 'api':
            recurso, id_str = parts[1], parts[2]
            try:
                rid = int(id_str)
            except ValueError:
                return self._send_json(400, {"error": "ID inválido"})

            if recurso == 'usuarios':
                count = self._db_execute_rowcount("DELETE FROM usuarios WHERE id = ?", (rid,))
                if count:
                    generate_excel()
                return self._send_json(200 if count else 404, {"deleted": bool(count)})
            if recurso == 'salas':
                count = self._db_execute_rowcount("DELETE FROM salas WHERE id = ?", (rid,))
                if count:
                    generate_excel()
                return self._send_json(200 if count else 404, {"deleted": bool(count)})
            if recurso == 'reservas':
                count = self._db_execute_rowcount("DELETE FROM reservas WHERE id = ?", (rid,))
                if count:
                    generate_excel()
                return self._send_json(200 if count else 404, {"deleted": bool(count)})

        return self._send_json(404, {"error": "Ruta no encontrada"})


# Cambiamos al directorio de frontend para servir los archivos estáticos
os.chdir("frontend")
server = HTTPServer(("localhost", 8000), RequestHandler)
print("Servidor iniciado en http://localhost:8000")
server.serve_forever()
